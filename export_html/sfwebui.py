# -*- coding: utf-8 -*-
# -----------------------------------------------------------------
# Name:         sfwebui
# Purpose:      User interface class for use with a web browser
#
# Author:       Steve Micallef <steve@binarypool.com>
#
# Created:      30/09/2012
# Copyright:    (c) Steve Micallef 2012
# License:      MIT
# -----------------------------------------------------------------
import csv
import html
import json
import logging
import multiprocessing as mp
import random
import string
import time
from copy import deepcopy
from io import BytesIO, StringIO
from operator import itemgetter
from pathlib import Path
import cherrypy
from cherrypy import _cperror
import os, time, re, traceback
from pathlib import Path

from mako.lookup import TemplateLookup
from mako.template import Template

import openpyxl

import secure

from sflib import SpiderFoot

from sfscan import startSpiderFootScanner

from spiderfoot import SpiderFootDb
from spiderfoot import SpiderFootHelpers
from spiderfoot import __version__
from spiderfoot.logger import logListenerSetup, logWorkerSetup

mp.set_start_method("spawn", force=True)


class SpiderFootWebUi:
    """SpiderFoot web interface."""

    lookup = TemplateLookup(directories=[''])
    defaultConfig = dict()
    config = dict()
    token = None
    docroot = ''

    def __init__(self: 'SpiderFootWebUi', web_config: dict, config: dict, loggingQueue: 'logging.handlers.QueueListener' = None) -> None:
        """Initialize web server.

        Args:
            web_config (dict): config settings for web interface (interface, port, root path)
            config (dict): SpiderFoot config
            loggingQueue: TBD

        Raises:
            TypeError: arg type is invalid
            ValueError: arg value is invalid
        """
        if not isinstance(config, dict):
            raise TypeError(f"config is {type(config)}; expected dict()")
        if not config:
            raise ValueError("config is empty")

        if not isinstance(web_config, dict):
            raise TypeError(f"web_config is {type(web_config)}; expected dict()")
        if not config:
            raise ValueError("web_config is empty")

        self.docroot = web_config.get('root', '/').rstrip('/')

        # 'config' supplied will be the defaults, let's supplement them
        # now with any configuration which may have previously been saved.
        self.defaultConfig = deepcopy(config)
        dbh = SpiderFootDb(self.defaultConfig, init=True)
        sf = SpiderFoot(self.defaultConfig)
        self.config = sf.configUnserialize(dbh.configGet(), self.defaultConfig)

        # Set up logging
        if loggingQueue is None:
            self.loggingQueue = mp.Queue()
            logListenerSetup(self.loggingQueue, self.config)
        else:
            self.loggingQueue = loggingQueue
        logWorkerSetup(self.loggingQueue)
        self.log = logging.getLogger(f"spiderfoot.{__name__}")

        cherrypy.config.update({
            'error_page.401': self.error_page_401,
            'error_page.404': self.error_page_404,
            'request.error_response': self.error_page
        })

        csp = (
            secure.ContentSecurityPolicy()
            .default_src("'self'")
            .script_src("'self'", "'unsafe-inline'", "blob:")
            .style_src("'self'", "'unsafe-inline'")
            .base_uri("'self'")
            .connect_src("'self'", "data:")
            .frame_src("'self'", 'data:')
            .img_src("'self'", "data:")
        )

        secure_headers = secure.Secure(
            server=secure.Server().set("server"),
            cache=secure.CacheControl().must_revalidate(),
            csp=csp,
            referrer=secure.ReferrerPolicy().no_referrer(),
        )

        cherrypy.config.update({
            "tools.response_headers.on": True,
            "tools.response_headers.headers": secure_headers.framework.cherrypy()
        })

    def error_page(self: 'SpiderFootWebUi') -> None:
        """Error page."""
        cherrypy.response.status = 500

        if self.config.get('_debug'):
            cherrypy.response.body = _cperror.get_error_page(status=500, traceback=_cperror.format_exc())
        else:
            cherrypy.response.body = b"<html><body>Error</body></html>"

    def error_page_401(self: 'SpiderFootWebUi', status: str, message: str, traceback: str, version: str) -> str:
        """Unauthorized access HTTP 401 error page.

        Args:
            status (str): HTTP response status code and message
            message (str): Error message
            traceback (str): Error stack trace
            version (str): CherryPy version

        Returns:
            str: HTML response
        """
        return ""

    def error_page_404(self: 'SpiderFootWebUi', status: str, message: str, traceback: str, version: str) -> str:
        """Not found error page 404.

        Args:
            status (str): HTTP response status code and message
            message (str): Error message
            traceback (str): Error stack trace
            version (str): CherryPy version

        Returns:
            str: HTTP response template
        """
        templ = Template(filename='spiderfoot/templates/error.tmpl', lookup=self.lookup)
        return templ.render(message='Not Found', docroot=self.docroot, status=status, version=__version__)

    def jsonify_error(self: 'SpiderFootWebUi', status: str, message: str) -> dict:
        """Jsonify error response.

        Args:
            status (str): HTTP response status code and message
            message (str): Error message

        Returns:
            dict: HTTP error response template
        """
        cherrypy.response.headers['Content-Type'] = 'application/json'
        cherrypy.response.status = status
        return {
            'error': {
                'http_status': status,
                'message': message,
            }
        }

    def error(self: 'SpiderFootWebUi', message: str) -> None:
        """Show generic error page with error message.

        Args:
            message (str): error message

        Returns:
            None
        """
        templ = Template(filename='spiderfoot/templates/error.tmpl', lookup=self.lookup)
        return templ.render(message=message, docroot=self.docroot, version=__version__)

    def cleanUserInput(self: 'SpiderFootWebUi', inputList: list) -> list:
        """Convert data to HTML entities; except quotes and ampersands.

        Args:
            inputList (list): list of strings to sanitize

        Returns:
            list: sanitized input

        Raises:
            TypeError: inputList type was invalid

        Todo:
            Review all uses of this function, then remove it.
            Use of this function is overloaded.
        """
        if not isinstance(inputList, list):
            raise TypeError(f"inputList is {type(inputList)}; expected list()")

        ret = list()

        for item in inputList:
            if not item:
                ret.append('')
                continue
            c = html.escape(item, True)

            # Decode '&' and '"' HTML entities
            c = c.replace("&amp;", "&").replace("&quot;", "\"")
            ret.append(c)

        return ret

    def searchBase(self: 'SpiderFootWebUi', id: str = None, eventType: str = None, value: str = None) -> list:
        """Search.

        Args:
            id (str): scan ID
            eventType (str): TBD
            value (str): TBD

        Returns:
            list: search results
        """
        retdata = []

        if not id and not eventType and not value:
            return retdata

        if not value:
            value = ''

        regex = ""
        if value.startswith("/") and value.endswith("/"):
            regex = value[1:len(value) - 1]
            value = ""

        value = value.replace('*', '%')
        if value in [None, ""] and regex in [None, ""]:
            value = "%"
            regex = ""

        dbh = SpiderFootDb(self.config)
        criteria = {
            'scan_id': id or '',
            'type': eventType or '',
            'value': value or '',
            'regex': regex or '',
        }

        try:
            data = dbh.search(criteria)
        except Exception:
            return retdata

        for row in data:
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            escapeddata = html.escape(row[1])
            escapedsrc = html.escape(row[2])
            retdata.append([lastseen, escapeddata, escapedsrc,
                            row[3], row[5], row[6], row[7], row[8], row[10],
                            row[11], row[4], row[13], row[14]])

        return retdata

    def buildExcel(self: 'SpiderFootWebUi', data: list, columnNames: list, sheetNameIndex: int = 0) -> str:
        """Convert supplied raw data into GEXF (Graph Exchange XML Format) format (e.g. for Gephi).

        Args:
            data (list): Scan result as list
            columnNames (list): column names
            sheetNameIndex (int): TBD

        Returns:
            str: Excel workbook
        """
        rowNums = dict()
        workbook = openpyxl.Workbook()
        defaultSheet = workbook.active
        columnNames.pop(sheetNameIndex)
        allowed_sheet_chars = string.ascii_uppercase + string.digits + '_'
        for row in data:
            sheetName = "".join([c for c in str(row.pop(sheetNameIndex)) if c.upper() in allowed_sheet_chars])
            try:
                sheet = workbook[sheetName]
            except KeyError:
                # Create sheet
                workbook.create_sheet(sheetName)
                sheet = workbook[sheetName]
                # Write headers
                for col_num, column_title in enumerate(columnNames, 1):
                    cell = sheet.cell(row=1, column=col_num)
                    cell.value = column_title
                rowNums[sheetName] = 2

            # Write row
            for col_num, cell_value in enumerate(row, 1):
                cell = sheet.cell(row=rowNums[sheetName], column=col_num)
                cell.value = cell_value

            rowNums[sheetName] += 1

        if rowNums:
            workbook.remove(defaultSheet)

        # Sort sheets alphabetically
        workbook._sheets.sort(key=lambda ws: ws.title)

        # Save workbook
        with BytesIO() as f:
            workbook.save(f)
            f.seek(0)
            return f.read()

    #
    # Melman Code
    #
    @cherrypy.expose
    def upload(self):
        templ = Template(filename='spiderfoot/templates/upload.tmpl', lookup=self.lookup)
        return templ.render(pageid='UPLOAD',
                            docroot=self.docroot,
                            version=__version__,
                            cherrypy=cherrypy,
                            show_success=False)

    @cherrypy.expose
    def uploadtxt(self, **params):
        try:
            part = cherrypy.request.params.get("txtfile")
            if not part or not hasattr(part, "file"):
                return "Nenhum arquivo recebido (param txtfile ausente)."

            raw_name = os.path.basename(getattr(part, "filename", "") or "")
            if not raw_name.lower().endswith(".txt"):
                return "Apenas arquivos .txt são aceitos."

            dest_dir = Path("./uploads")
            dest_dir.mkdir(parents=True, exist_ok=True)

            safe_name = re.sub(r"[^a-zA-Z0-9._-]", "_", raw_name)
            dest_path = dest_dir / safe_name

            with open(dest_path, "wb") as f:
                f.write(part.file.read())

            templ = Template(filename='spiderfoot/templates/upload.tmpl', lookup=self.lookup)
            return templ.render(pageid='UPLOAD',
                                docroot=self.docroot,
                                version=__version__,
                                cherrypy=cherrypy,
                                show_success=True)
        except Exception as e:
            cherrypy.log(f"[uploadtxt] Erro: {e}")
            cherrypy.log(traceback.format_exc())
            return f"Erro interno: {e}"

    @cherrypy.expose
    def uploadjson(self, **params):
        try:
            raw_json = cherrypy.request.params.get("jsondata")
            reportname = cherrypy.request.params.get("reportname") or "SpiderFoot"

            # Preferir arquivo se enviado
            part = cherrypy.request.params.get("jsonfile")
            if part and hasattr(part, "file"):
                try:
                    uploaded = part.file.read()
                    if isinstance(uploaded, bytes):
                        uploaded = uploaded.decode('utf-8', errors='replace')
                    raw_json = uploaded
                except Exception:
                    pass

            if not raw_json or not str(raw_json).strip():
                return "Nenhum JSON recebido (arquivo ou texto)."

            try:
                events = json.loads(raw_json)
                if not isinstance(events, list):
                    return "Formato inválido: JSON deve ser uma lista de eventos."
            except Exception as e:
                return f"JSON inválido: {e}"

            # Normalizar estruturas
            norm_events = []
            for ev in events:
                if not isinstance(ev, dict):
                    continue
                norm_events.append({
                    "data": str(ev.get("data", "")),
                    "event_type": ev.get("event_type", "UNKNOWN"),
                    "module": str(ev.get("module", "")),
                    "source_data": str(ev.get("source_data", "")),
                    "false_positive": ev.get("false_positive", 0),
                    "last_seen": ev.get("last_seen", ""),
                    "scan_name": ev.get("scan_name", reportname),
                    "scan_target": ev.get("scan_target", "")
                })

            html_content = self.gerar_html_dashboard_spiderfoot(norm_events, reportname, norm_events[0].get("scan_target", "") if norm_events else "")
            fname = (reportname or "SpiderFoot") + "-SpiderFoot.html"
            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "text/html; charset=utf-8"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return html_content.encode('utf-8')
        except Exception as e:
            cherrypy.log(f"[uploadjson] Erro: {e}")
            cherrypy.log(traceback.format_exc())
            return f"Erro interno: {e}"


    #
    # USER INTERFACE PAGES
    #

    @cherrypy.expose
    def scanexportlogs(self: 'SpiderFootWebUi', id: str, dialect: str = "excel") -> bytes:
        """Get scan log

        Args:
            id (str): scan ID
            dialect (str): CSV dialect (default: excel)

        Returns:
            bytes: scan logs in CSV format
        """
        dbh = SpiderFootDb(self.config)

        try:
            data = dbh.scanLogs(id, None, None, True)
        except Exception:
            return self.error("Scan ID not found.")

        if not data:
            return self.error("Scan ID not found.")

        fileobj = StringIO()
        parser = csv.writer(fileobj, dialect=dialect)
        parser.writerow(["Date", "Component", "Type", "Event", "Event ID"])
        for row in data:
            parser.writerow([
                time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000)),
                str(row[1]),
                str(row[2]),
                str(row[3]),
                row[4]
            ])

        cherrypy.response.headers['Content-Disposition'] = f"attachment; filename=SpiderFoot-{id}.log.csv"
        cherrypy.response.headers['Content-Type'] = "application/csv"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return fileobj.getvalue().encode('utf-8')

    @cherrypy.expose
    def scancorrelationsexport(self: 'SpiderFootWebUi', id: str, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get scan correlation data in CSV or Excel format.

        Args:
            id (str): scan ID
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        dbh = SpiderFootDb(self.config)

        try:
            scaninfo = dbh.scanInstanceGet(id)
            scan_name = scaninfo[0]
        except Exception:
            return json.dumps(["ERROR", "Could not retrieve info for scan."]).encode('utf-8')

        try:
            correlations = dbh.scanCorrelationList(id)
        except Exception:
            return json.dumps(["ERROR", "Could not retrieve correlations for scan."]).encode('utf-8')

        headings = ["Rule Name", "Correlation", "Risk", "Description"]

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in correlations:
                correlation = row[1]
                rule_name = row[2]
                rule_risk = row[3]
                rule_description = row[5]
                rows.append([rule_name, correlation, rule_risk, rule_description])

            if scan_name:
                fname = f"{scan_name}-SpiderFoot-correlations.xlxs"
            else:
                fname = "SpiderFoot-correlations.xlxs"

            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, headings, sheetNameIndex=0)

        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(headings)

            for row in correlations:
                correlation = row[1]
                rule_name = row[2]
                rule_risk = row[3]
                rule_description = row[5]
                parser.writerow([rule_name, correlation, rule_risk, rule_description])

            if scan_name:
                fname = f"{scan_name}-SpiderFoot-correlations.csv"
            else:
                fname = "SpiderFoot-correlations.csv"

            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scaneventresultexport(self: 'SpiderFootWebUi', id: str, type: str, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get scan event result data in CSV or Excel format

        Args:
            id (str): scan ID
            type (str): TBD
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        dbh = SpiderFootDb(self.config)
        data = dbh.scanResultEvent(id, type)

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                rows.append([lastseen, str(row[4]), str(row[3]), str(row[2]), row[13], datafield])

            fname = "SpiderFoot.xlsx"
            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, ["Updated", "Type", "Module", "Source",
                                   "F/P", "Data"], sheetNameIndex=1)

        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(["Updated", "Type", "Module", "Source", "F/P", "Data"])
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                parser.writerow([lastseen, str(row[4]), str(row[3]), str(row[2]), row[13], datafield])

            fname = "SpiderFoot.csv"
            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scaneventresultexportmulti(self: 'SpiderFootWebUi', ids: str, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get scan event result data in CSV or Excel format for multiple scans

        Args:
            ids (str): comma separated list of scan IDs
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        dbh = SpiderFootDb(self.config)
        scaninfo = dict()
        data = list()
        scan_name = ""

        for id in ids.split(','):
            scaninfo[id] = dbh.scanInstanceGet(id)
            if scaninfo[id] is None:
                continue
            scan_name = scaninfo[id][0]
            data = data + dbh.scanResultEvent(id)

        if not data:
            return None

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                rows.append([scaninfo[row[12]][0], lastseen, str(row[4]), str(row[3]),
                            str(row[2]), row[13], datafield])

            if len(ids.split(',')) > 1 or scan_name == "":
                fname = "SpiderFoot.xlsx"
            else:
                fname = scan_name + "-SpiderFoot.xlsx"

            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, ["Scan Name", "Updated", "Type", "Module",
                                   "Source", "F/P", "Data"], sheetNameIndex=2)

        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(["Scan Name", "Updated", "Type", "Module", "Source", "F/P", "Data"])
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                parser.writerow([scaninfo[row[12]][0], lastseen, str(row[4]), str(row[3]),
                                str(row[2]), row[13], datafield])

            if len(ids.split(',')) > 1 or scan_name == "":
                fname = "SpiderFoot.csv"
            else:
                fname = scan_name + "-SpiderFoot.csv"

            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scansearchresultexport(self: 'SpiderFootWebUi', id: str, eventType: str = None, value: str = None, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get search result data in CSV or Excel format

        Args:
            id (str): scan ID
            eventType (str): TBD
            value (str): TBD
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        data = self.searchBase(id, eventType, value)

        if not data:
            return None

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in data:
                if row[10] == "ROOT":
                    continue
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                rows.append([row[0], str(row[10]), str(row[3]), str(row[2]), row[11], datafield])
            cherrypy.response.headers['Content-Disposition'] = "attachment; filename=SpiderFoot.xlsx"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, ["Updated", "Type", "Module", "Source",
                                   "F/P", "Data"], sheetNameIndex=1)

        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(["Updated", "Type", "Module", "Source", "F/P", "Data"])
            for row in data:
                if row[10] == "ROOT":
                    continue
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                parser.writerow([row[0], str(row[10]), str(row[3]), str(row[2]), row[11], datafield])
            cherrypy.response.headers['Content-Disposition'] = "attachment; filename=SpiderFoot.csv"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scanexportjsonmulti(self: 'SpiderFootWebUi', ids: str) -> str:
        """Get scan event result data in JSON format for multiple scans.

        Args:
            ids (str): comma separated list of scan IDs

        Returns:
            str: results in JSON format
        """
        dbh = SpiderFootDb(self.config)
        scaninfo = list()
        scan_name = ""

        for id in ids.split(','):
            scan = dbh.scanInstanceGet(id)

            if scan is None:
                continue

            scan_name = scan[0]

            for row in dbh.scanResultEvent(id):
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                event_data = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                source_data = str(row[2])
                source_module = str(row[3])
                event_type = row[4]
                false_positive = row[13]

                if event_type == "ROOT":
                    continue

                scaninfo.append({
                    "data": event_data,
                    "event_type": event_type,
                    "module": source_module,
                    "source_data": source_data,
                    "false_positive": false_positive,
                    "last_seen": lastseen,
                    "scan_name": scan_name,
                    "scan_target": scan[1]
                })

        if len(ids.split(',')) > 1 or scan_name == "":
            fname = "SpiderFoot.json"
        else:
            fname = scan_name + "-SpiderFoot.json"

        cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return json.dumps(scaninfo).encode('utf-8')

    @cherrypy.expose
    def scanviz(self: 'SpiderFootWebUi', id: str, gexf: str = "0") -> str:
        """Export entities from scan results for visualising.

        Args:
            id (str): scan ID
            gexf (str): TBD

        Returns:
            str: GEXF data
        """
        if not id:
            return None

        dbh = SpiderFootDb(self.config)
        data = dbh.scanResultEvent(id, filterFp=True)
        scan = dbh.scanInstanceGet(id)

        if not scan:
            return None

        scan_name = scan[0]

        root = scan[1]

        if gexf == "0":
            return SpiderFootHelpers.buildGraphJson([root], data)

        if not scan_name:
            fname = "SpiderFoot.gexf"
        else:
            fname = scan_name + "SpiderFoot.gexf"

        cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
        cherrypy.response.headers['Content-Type'] = "application/gexf"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return SpiderFootHelpers.buildGraphGexf([root], "SpiderFoot Export", data)

    @cherrypy.expose
    def scanvizmulti(self: 'SpiderFootWebUi', ids: str, gexf: str = "1") -> str:
        """Export entities results from multiple scans in GEXF format.

        Args:
            ids (str): scan IDs
            gexf (str): TBD

        Returns:
            str: GEXF data
        """
        dbh = SpiderFootDb(self.config)
        data = list()
        roots = list()
        scan_name = ""

        if not ids:
            return None

        for id in ids.split(','):
            scan = dbh.scanInstanceGet(id)
            if not scan:
                continue
            data = data + dbh.scanResultEvent(id, filterFp=True)
            roots.append(scan[1])
            scan_name = scan[0]

        if not data:
            return None

        if gexf == "0":
            # Not implemented yet
            return None

        if len(ids.split(',')) > 1 or scan_name == "":
            fname = "SpiderFoot.gexf"
        else:
            fname = scan_name + "-SpiderFoot.gexf"

        cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
        cherrypy.response.headers['Content-Type'] = "application/gexf"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return SpiderFootHelpers.buildGraphGexf(roots, "SpiderFoot Export", data)

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanopts(self: 'SpiderFootWebUi', id: str) -> dict:
        """Return configuration used for the specified scan as JSON.

        Args:
            id: scan ID

        Returns:
            dict: scan options for the specified scan
        """
        dbh = SpiderFootDb(self.config)
        ret = dict()

        meta = dbh.scanInstanceGet(id)
        if not meta:
            return ret

        if meta[3] != 0:
            started = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(meta[3]))
        else:
            started = "Not yet"

        if meta[4] != 0:
            finished = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(meta[4]))
        else:
            finished = "Not yet"

        ret['meta'] = [meta[0], meta[1], meta[2], started, finished, meta[5]]
        ret['config'] = dbh.scanConfigGet(id)
        ret['configdesc'] = dict()
        for key in list(ret['config'].keys()):
            if ':' not in key:
                globaloptdescs = self.config['__globaloptdescs__']
                if globaloptdescs:
                    ret['configdesc'][key] = globaloptdescs.get(key, f"{key} (legacy)")
            else:
                [modName, modOpt] = key.split(':')
                if modName not in list(self.config['__modules__'].keys()):
                    continue

                if modOpt not in list(self.config['__modules__'][modName]['optdescs'].keys()):
                    continue

                ret['configdesc'][key] = self.config['__modules__'][modName]['optdescs'][modOpt]

        return ret

    @cherrypy.expose
    def rerunscan(self: 'SpiderFootWebUi', id: str) -> None:
        """Rerun a scan.

        Args:
            id (str): scan ID

        Returns:
            None

        Raises:
            HTTPRedirect: redirect to info page for new scan
        """
        # Snapshot the current configuration to be used by the scan
        cfg = deepcopy(self.config)
        modlist = list()
        dbh = SpiderFootDb(cfg)
        info = dbh.scanInstanceGet(id)

        if not info:
            return self.error("Invalid scan ID.")

        scanname = info[0]
        scantarget = info[1]

        scanconfig = dbh.scanConfigGet(id)
        if not scanconfig:
            return self.error(f"Error loading config from scan: {id}")

        modlist = scanconfig['_modulesenabled'].split(',')
        if "sfp__stor_stdout" in modlist:
            modlist.remove("sfp__stor_stdout")

        targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
        if not targetType:
            # It must then be a name, as a re-run scan should always have a clean
            # target. Put quotes around the target value and try to determine the
            # target type again.
            targetType = SpiderFootHelpers.targetTypeFromString(f'"{scantarget}"')

        if targetType not in ["HUMAN_NAME", "BITCOIN_ADDRESS"]:
            scantarget = scantarget.lower()

        # Start running a new scan
        scanId = SpiderFootHelpers.genScanInstanceId()
        try:
            p = mp.Process(target=startSpiderFootScanner, args=(self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg))
            p.daemon = True
            p.start()
        except Exception as e:
            self.log.error(f"[-] Scan [{scanId}] failed: {e}")
            return self.error(f"[-] Scan [{scanId}] failed: {e}")

        # Wait until the scan has initialized
        while dbh.scanInstanceGet(scanId) is None:
            self.log.info("Waiting for the scan to initialize...")
            time.sleep(1)

        raise cherrypy.HTTPRedirect(f"{self.docroot}/scaninfo?id={scanId}", status=302)

    @cherrypy.expose
    def rerunscanmulti(self: 'SpiderFootWebUi', ids: str) -> str:
        """Rerun scans.

        Args:
            ids (str): comma separated list of scan IDs

        Returns:
            str: Scan list page HTML
        """
        # Snapshot the current configuration to be used by the scan
        cfg = deepcopy(self.config)
        modlist = list()
        dbh = SpiderFootDb(cfg)

        for id in ids.split(","):
            info = dbh.scanInstanceGet(id)
            if not info:
                return self.error("Invalid scan ID.")

            scanconfig = dbh.scanConfigGet(id)
            scanname = info[0]
            scantarget = info[1]
            targetType = None

            if len(scanconfig) == 0:
                return self.error("Something went wrong internally.")

            modlist = scanconfig['_modulesenabled'].split(',')
            if "sfp__stor_stdout" in modlist:
                modlist.remove("sfp__stor_stdout")

            targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
            if targetType is None:
                # Should never be triggered for a re-run scan..
                return self.error("Invalid target type. Could not recognize it as a target SpiderFoot supports.")

            # Start running a new scan
            scanId = SpiderFootHelpers.genScanInstanceId()
            try:
                p = mp.Process(target=startSpiderFootScanner, args=(self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg))
                p.daemon = True
                p.start()
            except Exception as e:
                self.log.error(f"[-] Scan [{scanId}] failed: {e}")
                return self.error(f"[-] Scan [{scanId}] failed: {e}")

            # Wait until the scan has initialized
            while dbh.scanInstanceGet(scanId) is None:
                self.log.info("Waiting for the scan to initialize...")
                time.sleep(1)

        templ = Template(filename='spiderfoot/templates/scanlist.tmpl', lookup=self.lookup)
        return templ.render(rerunscans=True, docroot=self.docroot, pageid="SCANLIST", version=__version__)

    @cherrypy.expose
    def newscan(self: 'SpiderFootWebUi') -> str:
        """Configure a new scan.

        Returns:
            str: New scan page HTML
        """
        dbh = SpiderFootDb(self.config)
        types = dbh.eventTypes()
        templ = Template(filename='spiderfoot/templates/newscan.tmpl', lookup=self.lookup)
        return templ.render(pageid='NEWSCAN', types=types, docroot=self.docroot,
                            modules=self.config['__modules__'], scanname="",
                            selectedmods="", scantarget="", version=__version__)

    @cherrypy.expose
    def clonescan(self: 'SpiderFootWebUi', id: str) -> str:
        """Clone an existing scan (pre-selected options in the newscan page).

        Args:
            id (str): scan ID to clone

        Returns:
            str: New scan page HTML pre-populated with options from cloned scan.
        """
        dbh = SpiderFootDb(self.config)
        types = dbh.eventTypes()
        info = dbh.scanInstanceGet(id)

        if not info:
            return self.error("Invalid scan ID.")

        scanconfig = dbh.scanConfigGet(id)
        scanname = info[0]
        scantarget = info[1]
        targetType = None

        if scanname == "" or scantarget == "" or len(scanconfig) == 0:
            return self.error("Something went wrong internally.")

        targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
        if targetType is None:
            # It must be a name, so wrap quotes around it
            scantarget = "&quot;" + scantarget + "&quot;"

        modlist = scanconfig['_modulesenabled'].split(',')

        templ = Template(filename='spiderfoot/templates/newscan.tmpl', lookup=self.lookup)
        return templ.render(pageid='NEWSCAN', types=types, docroot=self.docroot,
                            modules=self.config['__modules__'], selectedmods=modlist,
                            scanname=str(scanname),
                            scantarget=str(scantarget), version=__version__)

    @cherrypy.expose
    def index(self: 'SpiderFootWebUi') -> str:
        """Show scan list page.

        Returns:
            str: Scan list page HTML
        """
        templ = Template(filename='spiderfoot/templates/scanlist.tmpl', lookup=self.lookup)
        return templ.render(pageid='SCANLIST', docroot=self.docroot, version=__version__)

    @cherrypy.expose
    def scaninfo(self: 'SpiderFootWebUi', id: str) -> str:
        """Information about a selected scan.

        Args:
            id (str): scan id

        Returns:
            str: scan info page HTML
        """
        dbh = SpiderFootDb(self.config)
        res = dbh.scanInstanceGet(id)
        if res is None:
            return self.error("Scan ID not found.")

        templ = Template(filename='spiderfoot/templates/scaninfo.tmpl', lookup=self.lookup, input_encoding='utf-8')
        return templ.render(id=id, name=html.escape(res[0]), status=res[5], docroot=self.docroot, version=__version__,
                            pageid="SCANLIST")

    @cherrypy.expose
    def opts(self: 'SpiderFootWebUi', updated: str = None) -> str:
        """Show module and global settings page.

        Args:
            updated (str): scan options were updated successfully

        Returns:
            str: scan options page HTML
        """
        templ = Template(filename='spiderfoot/templates/opts.tmpl', lookup=self.lookup)
        self.token = random.SystemRandom().randint(0, 99999999)
        return templ.render(opts=self.config, pageid='SETTINGS', token=self.token, version=__version__,
                            updated=updated, docroot=self.docroot)

    @cherrypy.expose
    def optsexport(self: 'SpiderFootWebUi', pattern: str = None) -> str:
        """Export configuration.

        Args:
            pattern (str): TBD

        Returns:
            str: Configuration settings
        """
        sf = SpiderFoot(self.config)
        conf = sf.configSerialize(self.config)
        content = ""

        for opt in sorted(conf):
            if ":_" in opt or opt.startswith("_"):
                continue

            if pattern:
                if pattern in opt:
                    content += f"{opt}={conf[opt]}\n"
            else:
                content += f"{opt}={conf[opt]}\n"

        cherrypy.response.headers['Content-Disposition'] = 'attachment; filename="SpiderFoot.cfg"'
        cherrypy.response.headers['Content-Type'] = "text/plain"
        return content

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def optsraw(self: 'SpiderFootWebUi') -> str:
        """Return global and module settings as json.

        Returns:
            str: settings as JSON
        """
        ret = dict()
        self.token = random.SystemRandom().randint(0, 99999999)
        for opt in self.config:
            if not opt.startswith('__'):
                ret["global." + opt] = self.config[opt]
                continue

            if opt == '__modules__':
                for mod in sorted(self.config['__modules__'].keys()):
                    for mo in sorted(self.config['__modules__'][mod]['opts'].keys()):
                        if mo.startswith("_"):
                            continue
                        ret["module." + mod + "." + mo] = self.config['__modules__'][mod]['opts'][mo]

        return ['SUCCESS', {'token': self.token, 'data': ret}]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scandelete(self: 'SpiderFootWebUi', id: str) -> str:
        """Delete scan(s).

        Args:
            id (str): comma separated list of scan IDs

        Returns:
            str: JSON response
        """
        if not id:
            return self.jsonify_error('404', "No scan specified")

        dbh = SpiderFootDb(self.config)
        ids = id.split(',')

        for scan_id in ids:
            res = dbh.scanInstanceGet(scan_id)
            if not res:
                return self.jsonify_error('404', f"Scan {scan_id} does not exist")

            if res[5] in ["RUNNING", "STARTING", "STARTED"]:
                return self.jsonify_error('400', f"Scan {scan_id} is {res[5]}. You cannot delete running scans.")

        for scan_id in ids:
            dbh.scanInstanceDelete(scan_id)

        return ""

    @cherrypy.expose
    def savesettings(self: 'SpiderFootWebUi', allopts: str, token: str, configFile: 'cherrypy._cpreqbody.Part' = None) -> None:
        """Save settings, also used to completely reset them to default.

        Args:
            allopts: TBD
            token (str): CSRF token
            configFile (cherrypy._cpreqbody.Part): TBD

        Returns:
            None

        Raises:
            HTTPRedirect: redirect to scan settings
        """
        if str(token) != str(self.token):
            return self.error(f"Invalid token ({token})")

        # configFile seems to get set even if a file isn't uploaded
        if configFile and configFile.file:
            try:
                contents = configFile.file.read()

                if isinstance(contents, bytes):
                    contents = contents.decode('utf-8')

                tmp = dict()
                for line in contents.split("\n"):
                    if "=" not in line:
                        continue

                    opt_array = line.strip().split("=")
                    if len(opt_array) == 1:
                        opt_array[1] = ""

                    tmp[opt_array[0]] = '='.join(opt_array[1:])

                allopts = json.dumps(tmp).encode('utf-8')
            except Exception as e:
                return self.error(f"Failed to parse input file. Was it generated from SpiderFoot? ({e})")

        # Reset config to default
        if allopts == "RESET":
            if self.reset_settings():
                raise cherrypy.HTTPRedirect(f"{self.docroot}/opts?updated=1")
            return self.error("Failed to reset settings")

        # Save settings
        try:
            dbh = SpiderFootDb(self.config)
            useropts = json.loads(allopts)
            cleanopts = dict()
            for opt in list(useropts.keys()):
                cleanopts[opt] = self.cleanUserInput([useropts[opt]])[0]

            currentopts = deepcopy(self.config)

            # Make a new config where the user options override
            # the current system config.
            sf = SpiderFoot(self.config)
            self.config = sf.configUnserialize(cleanopts, currentopts)
            dbh.configSet(sf.configSerialize(self.config))
        except Exception as e:
            return self.error(f"Processing one or more of your inputs failed: {e}")

        raise cherrypy.HTTPRedirect(f"{self.docroot}/opts?updated=1")

    @cherrypy.expose
    def savesettingsraw(self: 'SpiderFootWebUi', allopts: str, token: str) -> str:
        """Save settings, also used to completely reset them to default.

        Args:
            allopts: TBD
            token (str): CSRF token

        Returns:
            str: save success as JSON
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        if str(token) != str(self.token):
            return json.dumps(["ERROR", f"Invalid token ({token})."]).encode('utf-8')

        # Reset config to default
        if allopts == "RESET":
            if self.reset_settings():
                return json.dumps(["SUCCESS", ""]).encode('utf-8')
            return json.dumps(["ERROR", "Failed to reset settings"]).encode('utf-8')

        # Save settings
        try:
            dbh = SpiderFootDb(self.config)
            useropts = json.loads(allopts)
            cleanopts = dict()
            for opt in list(useropts.keys()):
                cleanopts[opt] = self.cleanUserInput([useropts[opt]])[0]

            currentopts = deepcopy(self.config)

            # Make a new config where the user options override
            # the current system config.
            sf = SpiderFoot(self.config)
            self.config = sf.configUnserialize(cleanopts, currentopts)
            dbh.configSet(sf.configSerialize(self.config))
        except Exception as e:
            return json.dumps(["ERROR", f"Processing one or more of your inputs failed: {e}"]).encode('utf-8')

        return json.dumps(["SUCCESS", ""]).encode('utf-8')

    def reset_settings(self: 'SpiderFootWebUi') -> bool:
        """Reset settings to default.

        Returns:
            bool: success
        """
        try:
            dbh = SpiderFootDb(self.config)
            dbh.configClear()  # Clear it in the DB
            self.config = deepcopy(self.defaultConfig)  # Clear in memory
        except Exception:
            return False

        return True

    @cherrypy.expose
    def resultsetfp(self: 'SpiderFootWebUi', id: str, resultids: str, fp: str) -> str:
        """Set a bunch of results (hashes) as false positive.

        Args:
            id (str): scan ID
            resultids (str): comma separated list of result IDs
            fp (str): 0 or 1

        Returns:
            str: set false positive status as JSON
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = SpiderFootDb(self.config)

        if fp not in ["0", "1"]:
            return json.dumps(["ERROR", "No FP flag set or not set correctly."]).encode('utf-8')

        try:
            ids = json.loads(resultids)
        except Exception:
            return json.dumps(["ERROR", "No IDs supplied."]).encode('utf-8')

        # Cannot set FPs if a scan is not completed
        status = dbh.scanInstanceGet(id)
        if not status:
            return self.error(f"Invalid scan ID: {id}")

        if status[5] not in ["ABORTED", "FINISHED", "ERROR-FAILED"]:
            return json.dumps([
                "WARNING",
                "Scan must be in a finished state when setting False Positives."
            ]).encode('utf-8')

        # Make sure the user doesn't set something as non-FP when the
        # parent is set as an FP.
        if fp == "0":
            data = dbh.scanElementSourcesDirect(id, ids)
            for row in data:
                if str(row[14]) == "1":
                    return json.dumps([
                        "WARNING",
                        f"Cannot unset element {id} as False Positive if a parent element is still False Positive."
                    ]).encode('utf-8')

        # Set all the children as FPs too.. it's only logical afterall, right?
        childs = dbh.scanElementChildrenAll(id, ids)
        allIds = ids + childs

        ret = dbh.scanResultsUpdateFP(id, allIds, fp)
        if ret:
            return json.dumps(["SUCCESS", ""]).encode('utf-8')

        return json.dumps(["ERROR", "Exception encountered."]).encode('utf-8')

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def eventtypes(self: 'SpiderFootWebUi') -> list:
        """List all event types.

        Returns:
            list: list of event types
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = SpiderFootDb(self.config)
        types = dbh.eventTypes()
        ret = list()

        for r in types:
            ret.append([r[1], r[0]])

        return sorted(ret, key=itemgetter(0))

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def modules(self: 'SpiderFootWebUi') -> list:
        """List all modules.

        Returns:
            list: list of modules
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        ret = list()

        modinfo = list(self.config['__modules__'].keys())
        if not modinfo:
            return ret

        modinfo.sort()

        for m in modinfo:
            if "__" in m:
                continue
            ret.append({'name': m, 'descr': self.config['__modules__'][m]['descr']})

        return ret

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def correlationrules(self: 'SpiderFootWebUi') -> list:
        """List all correlation rules.

        Returns:
            list: list of correlation rules
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        ret = list()

        rules = self.config['__correlationrules__']
        if not rules:
            return ret

        for r in rules:
            ret.append({
                'id': r['id'],
                'name': r['meta']['name'],
                'descr': r['meta']['description'],
                'risk': r['meta']['risk'],
            })

        return ret

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def ping(self: 'SpiderFootWebUi') -> list:
        """For the CLI to test connectivity to this server.

        Returns:
            list: SpiderFoot version as JSON
        """
        return ["SUCCESS", __version__]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def query(self: 'SpiderFootWebUi', query: str) -> str:
        """For the CLI to run queries against the database.

        Args:
            query (str): SQL query

        Returns:
            str: query results as JSON
        """
        dbh = SpiderFootDb(self.config)

        if not query:
            return self.jsonify_error('400', "Invalid query.")

        if not query.lower().startswith("select"):
            return self.jsonify_error('400', "Non-SELECTs are unpredictable and not recommended.")

        try:
            ret = dbh.dbh.execute(query)
            data = ret.fetchall()
            columnNames = [c[0] for c in dbh.dbh.description]
            return [dict(zip(columnNames, row)) for row in data]
        except Exception as e:
            return self.jsonify_error('500', str(e))

    @cherrypy.expose
    def startscan(self: 'SpiderFootWebUi', scanname: str, scantarget: str, modulelist: str, typelist: str, usecase: str) -> str:
        """Initiate a scan.

        Args:
            scanname (str): scan name
            scantarget (str): scan target
            modulelist (str): comma separated list of modules to use
            typelist (str): selected modules based on produced event data types
            usecase (str): selected module group (passive, investigate, footprint, all)

        Returns:
            str: start scan status as JSON

        Raises:
            HTTPRedirect: redirect to new scan info page
        """
        scanname = self.cleanUserInput([scanname])[0]
        scantarget = self.cleanUserInput([scantarget])[0]

        if not scanname:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Incorrect usage: scan name was not specified."]).encode('utf-8')

            return self.error("Invalid request: scan name was not specified.")

        if not scantarget:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Incorrect usage: scan target was not specified."]).encode('utf-8')

            return self.error("Invalid request: scan target was not specified.")

        if not typelist and not modulelist and not usecase:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Incorrect usage: no modules specified for scan."]).encode('utf-8')

            return self.error("Invalid request: no modules specified for scan.")

        targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
        if targetType is None:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Unrecognised target type."]).encode('utf-8')

            return self.error("Invalid target type. Could not recognize it as a target SpiderFoot supports.")

        # Swap the globalscantable for the database handler
        dbh = SpiderFootDb(self.config)

        # Snapshot the current configuration to be used by the scan
        cfg = deepcopy(self.config)
        sf = SpiderFoot(cfg)

        modlist = list()

        # User selected modules
        if modulelist:
            modlist = modulelist.replace('module_', '').split(',')

        # User selected types
        if len(modlist) == 0 and typelist:
            typesx = typelist.replace('type_', '').split(',')

            # 1. Find all modules that produce the requested types
            modlist = sf.modulesProducing(typesx)
            newmods = deepcopy(modlist)
            newmodcpy = deepcopy(newmods)

            # 2. For each type those modules consume, get modules producing
            while len(newmodcpy) > 0:
                for etype in sf.eventsToModules(newmodcpy):
                    xmods = sf.modulesProducing([etype])
                    for mod in xmods:
                        if mod not in modlist:
                            modlist.append(mod)
                            newmods.append(mod)
                newmodcpy = deepcopy(newmods)
                newmods = list()

        # User selected a use case
        if len(modlist) == 0 and usecase:
            for mod in self.config['__modules__']:
                if usecase == 'all' or usecase in self.config['__modules__'][mod]['group']:
                    modlist.append(mod)

        # If we somehow got all the way through to here and still don't have any modules selected
        if not modlist:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Incorrect usage: no modules specified for scan."]).encode('utf-8')

            return self.error("Invalid request: no modules specified for scan.")

        # Add our mandatory storage module
        if "sfp__stor_db" not in modlist:
            modlist.append("sfp__stor_db")
        modlist.sort()

        # Delete the stdout module in case it crept in
        if "sfp__stor_stdout" in modlist:
            modlist.remove("sfp__stor_stdout")

        # Start running a new scan
        if targetType in ["HUMAN_NAME", "USERNAME", "BITCOIN_ADDRESS"]:
            scantarget = scantarget.replace("\"", "")
        else:
            scantarget = scantarget.lower()

        # Start running a new scan
        scanId = SpiderFootHelpers.genScanInstanceId()
        try:
            p = mp.Process(target=startSpiderFootScanner, args=(self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg))
            p.daemon = True
            p.start()
        except Exception as e:
            self.log.error(f"[-] Scan [{scanId}] failed: {e}")
            return self.error(f"[-] Scan [{scanId}] failed: {e}")

        # Wait until the scan has initialized
        # Check the database for the scan status results
        while dbh.scanInstanceGet(scanId) is None:
            self.log.info("Waiting for the scan to initialize...")
            time.sleep(1)

        if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
            cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
            return json.dumps(["SUCCESS", scanId]).encode('utf-8')

        raise cherrypy.HTTPRedirect(f"{self.docroot}/scaninfo?id={scanId}")

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def stopscan(self: 'SpiderFootWebUi', id: str) -> str:
        """Stop a scan.

        Args:
            id (str): comma separated list of scan IDs

        Returns:
            str: JSON response
        """
        if not id:
            return self.jsonify_error('404', "No scan specified")

        dbh = SpiderFootDb(self.config)
        ids = id.split(',')

        for scan_id in ids:
            res = dbh.scanInstanceGet(scan_id)
            if not res:
                return self.jsonify_error('404', f"Scan {scan_id} does not exist")

            scan_status = res[5]

            if scan_status == "FINISHED":
                return self.jsonify_error('400', f"Scan {scan_id} has already finished.")

            if scan_status == "ABORTED":
                return self.jsonify_error('400', f"Scan {scan_id} has already aborted.")

            if scan_status != "RUNNING" and scan_status != "STARTING":
                return self.jsonify_error('400', f"The running scan is currently in the state '{scan_status}', please try again later or restart SpiderFoot.")

        for scan_id in ids:
            dbh.scanInstanceSet(scan_id, status="ABORT-REQUESTED")

        return ""

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def vacuum(self):
        dbh = SpiderFootDb(self.config)
        try:
            if dbh.vacuumDB():
                return json.dumps(["SUCCESS", ""]).encode('utf-8')
            return json.dumps(["ERROR", "Vacuuming the database failed"]).encode('utf-8')
        except Exception as e:
            return json.dumps(["ERROR", f"Vacuuming the database failed: {e}"]).encode('utf-8')

    #
    # DATA PROVIDERS
    #

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanlog(self: 'SpiderFootWebUi', id: str, limit: str = None, rowId: str = None, reverse: str = None) -> list:
        """Scan log data.

        Args:
            id (str): scan ID
            limit (str): TBD
            rowId (str): TBD
            reverse (str): TBD

        Returns:
            list: scan log
        """
        dbh = SpiderFootDb(self.config)
        retdata = []

        try:
            data = dbh.scanLogs(id, limit, rowId, reverse)
        except Exception:
            return retdata

        for row in data:
            generated = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000))
            retdata.append([generated, row[1], row[2], html.escape(row[3]), row[4]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanerrors(self: 'SpiderFootWebUi', id: str, limit: str = None) -> list:
        """Scan error data.

        Args:
            id (str): scan ID
            limit (str): limit number of results

        Returns:
            list: scan errors
        """
        dbh = SpiderFootDb(self.config)
        retdata = []

        try:
            data = dbh.scanErrors(id, limit)
        except Exception:
            return retdata

        for row in data:
            generated = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000))
            retdata.append([generated, row[1], html.escape(str(row[2]))])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanlist(self: 'SpiderFootWebUi') -> list:
        """Produce a list of scans.

        Returns:
            list: scan list
        """
        dbh = SpiderFootDb(self.config)
        data = dbh.scanInstanceList()
        retdata = []

        for row in data:
            created = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[3]))
            riskmatrix = {
                "HIGH": 0,
                "MEDIUM": 0,
                "LOW": 0,
                "INFO": 0
            }
            correlations = dbh.scanCorrelationSummary(row[0], by="risk")
            if correlations:
                for c in correlations:
                    riskmatrix[c[0]] = c[1]

            if row[4] == 0:
                started = "Not yet"
            else:
                started = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[4]))

            if row[5] == 0:
                finished = "Not yet"
            else:
                finished = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[5]))

            retdata.append([row[0], row[1], row[2], created, started, finished, row[6], row[7], riskmatrix])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanstatus(self: 'SpiderFootWebUi', id: str) -> list:
        """Show basic information about a scan, including status and number of each event type.

        Args:
            id (str): scan ID

        Returns:
            list: scan status
        """
        dbh = SpiderFootDb(self.config)
        data = dbh.scanInstanceGet(id)

        if not data:
            return []

        created = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[2]))
        started = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[3]))
        ended = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[4]))
        riskmatrix = {
            "HIGH": 0,
            "MEDIUM": 0,
            "LOW": 0,
            "INFO": 0
        }
        correlations = dbh.scanCorrelationSummary(id, by="risk")
        if correlations:
            for c in correlations:
                riskmatrix[c[0]] = c[1]

        return [data[0], data[1], created, started, ended, data[5], riskmatrix]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scansummary(self: 'SpiderFootWebUi', id: str, by: str) -> list:
        """Summary of scan results.

        Args:
            id (str): scan ID
            by (str): filter by type

        Returns:
            list: scan summary
        """
        retdata = []

        dbh = SpiderFootDb(self.config)

        try:
            scandata = dbh.scanResultSummary(id, by)
        except Exception:
            return retdata

        try:
            statusdata = dbh.scanInstanceGet(id)
        except Exception:
            return retdata

        for row in scandata:
            if row[0] == "ROOT":
                continue
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[2]))
            retdata.append([row[0], row[1], lastseen, row[3], row[4], statusdata[5]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scancorrelations(self: 'SpiderFootWebUi', id: str) -> list:
        """Correlation results from a scan.

        Args:
            id (str): scan ID

        Returns:
            list: correlation result list
        """
        retdata = []

        dbh = SpiderFootDb(self.config)

        try:
            corrdata = dbh.scanCorrelationList(id)
        except Exception:
            return retdata

        for row in corrdata:
            retdata.append([row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scaneventresults(self: 'SpiderFootWebUi', id: str, eventType: str = None, filterfp: bool = False, correlationId: str = None) -> list:
        """Return all event results for a scan as JSON.

        Args:
            id (str): scan ID
            eventType (str): filter by event type
            filterfp (bool): remove false positives from search results
            correlationId (str): filter by events associated with a correlation

        Returns:
            list: scan results
        """
        retdata = []

        dbh = SpiderFootDb(self.config)

        if not eventType:
            eventType = 'ALL'

        try:
            data = dbh.scanResultEvent(id, eventType, filterfp, correlationId=correlationId)
        except Exception:
            return retdata

        for row in data:
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            retdata.append([
                lastseen,
                html.escape(row[1]),
                html.escape(row[2]),
                row[3],
                row[5],
                row[6],
                row[7],
                row[8],
                row[13],
                row[14],
                row[4]
            ])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scaneventresultsunique(self: 'SpiderFootWebUi', id: str, eventType: str, filterfp: bool = False) -> list:
        """Return unique event results for a scan as JSON.

        Args:
            id (str): filter search results by scan ID
            eventType (str): filter search results by event type
            filterfp (bool): remove false positives from search results

        Returns:
            list: unique search results
        """
        dbh = SpiderFootDb(self.config)
        retdata = []

        try:
            data = dbh.scanResultEventUnique(id, eventType, filterfp)
        except Exception:
            return retdata

        for row in data:
            escaped = html.escape(row[0])
            retdata.append([escaped, row[1], row[2]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def search(self: 'SpiderFootWebUi', id: str = None, eventType: str = None, value: str = None) -> list:
        """Search scans.

        Args:
            id (str): filter search results by scan ID
            eventType (str): filter search results by event type
            value (str): filter search results by event value

        Returns:
            list: search results
        """
        try:
            return self.searchBase(id, eventType, value)
        except Exception:
            return []

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanhistory(self: 'SpiderFootWebUi', id: str) -> list:
        """Historical data for a scan.

        Args:
            id (str): scan ID

        Returns:
            list: scan history
        """
        if not id:
            return self.jsonify_error('404', "No scan specified")

        dbh = SpiderFootDb(self.config)

        try:
            return dbh.scanResultHistory(id)
        except Exception:
            return []

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanelementtypediscovery(self: 'SpiderFootWebUi', id: str, eventType: str) -> dict:
        """Scan element type discovery.

        Args:
            id (str): scan ID
            eventType (str): filter by event type

        Returns:
            dict
        """
        dbh = SpiderFootDb(self.config)
        pc = dict()
        datamap = dict()
        retdata = dict()

        # Get the events we will be tracing back from
        try:
            leafSet = dbh.scanResultEvent(id, eventType)
            [datamap, pc] = dbh.scanElementSourcesAll(id, leafSet)
        except Exception:
            return retdata

        # Delete the ROOT key as it adds no value from a viz perspective
        del pc['ROOT']
        retdata['tree'] = SpiderFootHelpers.dataParentChildToTree(pc)
        retdata['data'] = datamap

        return retdata

    # -----------------------------
    # Export: HTML Dashboard
    # -----------------------------
    @cherrypy.expose
    def scanexporthtml(self: 'SpiderFootWebUi', id: str) -> bytes:
        """Exporta resultados do scan como um Dashboard HTML único (arquivo).

        Args:
            id (str): scan ID

        Returns:
            bytes: conteúdo HTML do dashboard
        """
        dbh = SpiderFootDb(self.config)

        scan = dbh.scanInstanceGet(id)
        if not scan:
            return self.error("Scan ID not found.")

        scan_name = scan[0]
        scan_target = scan[1]

        # Coletar eventos e normalizar em dicionários
        events = []
        for row in dbh.scanResultEvent(id):
            # row indices per SpiderFoot: [0]=lastseen, [1]=data, [2]=sourceData, [3]=module, [4]=type, ... [13]=fp
            if row[4] == "ROOT":
                continue
            lastseen_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            events.append({
                "data": str(row[1]).replace("<SFURL>", "").replace("</SFURL>", ""),
                "event_type": row[4],
                "module": str(row[3]),
                "source_data": str(row[2]),
                "false_positive": row[13],
                "last_seen": lastseen_str,
                "scan_name": scan_name,
                "scan_target": scan_target
            })

        html_content = self.gerar_html_dashboard_spiderfoot(events, scan_name, scan_target)

        fname = (scan_name or "SpiderFoot") + "-SpiderFoot.html"
        cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
        cherrypy.response.headers['Content-Type'] = "text/html; charset=utf-8"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return html_content.encode('utf-8')

    # -----------------------------
    # HTML generator (Dashboard)
    # -----------------------------
    def gerar_html_dashboard_spiderfoot(self, events: list, scan_name: str, scan_target: str) -> str:
        """Gera um dashboard HTML a partir da lista normalizada de eventos.

        - Inclui abas: Resumo, Vulnerabilidades (CVE) e Detalhes
        - Na aba de Vulnerabilidades, a descrição tem botão Expandir/Recolher
        """
        # KPI CVEs por severidade
        cve_types = {
            "CRITICAL": "VULNERABILITY_CVE_CRITICAL",
            "HIGH": "VULNERABILITY_CVE_HIGH",
            "MEDIUM": "VULNERABILITY_CVE_MEDIUM",
            "LOW": "VULNERABILITY_CVE_LOW"
        }
        kpi = {"CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0}

        for e in events:
            et = e.get("event_type")
            if et == cve_types["CRITICAL"]:
                kpi["CRITICAL"] += 1
            elif et == cve_types["HIGH"]:
                kpi["HIGH"] += 1
            elif et == cve_types["MEDIUM"]:
                kpi["MEDIUM"] += 1
            elif et == cve_types["LOW"]:
                kpi["LOW"] += 1

        # Top 10 tipos de eventos (para o gráfico resumo)
        try:
            from collections import Counter
        except Exception:
            Counter = None  # graceful fallback
        labels_js = "[]"
        values_js = "[]"
        if events:
            if Counter is not None:
                counts = Counter([str(e.get("event_type", "UNKNOWN")) for e in events])
                top = counts.most_common(10)
            else:
                tmp = dict()
                for e in events:
                    etp = str(e.get("event_type", "UNKNOWN"))
                    tmp[etp] = tmp.get(etp, 0) + 1
                top = sorted(tmp.items(), key=lambda x: x[1], reverse=True)[:10]

            top_labels = [t[0] for t in top]
            top_values = [t[1] for t in top]
            labels_js = json.dumps(top_labels)
            values_js = json.dumps(top_values)

        resumo_html = self._build_resumo_section(kpi, len(events), scan_name, scan_target)
        vulns_html = self._build_cve_section(events)
        detalhes_html, detail_slugs = self._build_details_section(events)
        detail_slugs_js = json.dumps(detail_slugs)

        page = f"""
<!doctype html>
<html lang="pt-BR">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>SpiderFoot Dashboard - {html.escape(scan_name or 'Scan')}</title>
  <style>
    :root {{
      --ti-orange: #f57c00;
      --ti-orange-dark: #e65100;
      --ti-orange-light: #ff9800;
      --bg: #0f0f0f;
      --surface: #1b1b1b;
      --text: #ececec;
      --muted: #bdbdbd;
      --stroke: #2a2a2a;
      --accent: var(--ti-orange);
    }}
    body {{ font-family: "Segoe UI", Arial, Helvetica, sans-serif; margin: 0; padding: 0; color: var(--text); background: radial-gradient(1200px 800px at 10% -10%, #1b1b1b, #0f0f0f); }}
    header {{ padding: 18px 22px; background: linear-gradient(90deg, var(--ti-orange-dark), var(--ti-orange)); color: #fff; display: flex; align-items: center; gap: 12px; box-shadow: 0 2px 0 rgba(0,0,0,.2); }}
    header h1 {{ margin: 0; font-size: 18px; font-weight: 600; letter-spacing: .3px; }}
    .ti-lock {{ width: 28px; height: 28px; fill: #fff; filter: drop-shadow(0 1px 1px rgba(0,0,0,.3)); }}
    .container {{ max-width: 1240px; margin: 0 auto; padding: 0 18px 24px; }}
    .tabs {{ display: flex; gap: 10px; padding: 14px 0; position: sticky; top: 0; z-index: 10; background: rgba(15,15,15,.75); backdrop-filter: blur(6px); border-bottom: 1px solid var(--stroke); }}
    .tab-btn {{ border: 1px solid var(--stroke); background: linear-gradient(180deg, #1f1f1f, #181818); color: var(--text); padding: 10px 14px; cursor: pointer; border-radius: 10px; transition: all .18s ease; box-shadow: 0 1px 0 rgba(255,255,255,.04) inset; }}
    .tab-btn:hover {{ border-color: var(--ti-orange); color: var(--ti-orange-light); transform: translateY(-1px); }}
    .tab-btn.active {{ background: linear-gradient(180deg, #ffa726, #fb8c00); color: #111; border-color: var(--accent); font-weight: 700; box-shadow: 0 6px 22px rgba(251,140,0,.25); }}
    .tab-content {{ display: none; padding: 16px; }}
    .tab-content.active {{ display: block; }}
    .kpis {{ display: grid; grid-template-columns: repeat(5, minmax(160px, 1fr)); gap: 12px; margin-bottom: 16px; }}
    .kpi {{ border: 1px solid var(--stroke); border-radius: 14px; padding: 14px; background: linear-gradient(180deg, #171717, #131313); box-shadow: 0 1px 0 rgba(255,255,255,.04) inset, 0 10px 24px rgba(0,0,0,.25); }}
    .kpi-title {{ font-size: 12px; color: var(--muted); }}
    .kpi-value {{ font-size: 22px; font-weight: 700; margin-top: 6px; color: var(--ti-orange-light); }}
    .card {{ border: 1px solid var(--stroke); background: linear-gradient(180deg, #171717, #131313); border-radius: 14px; padding: 16px; margin: 12px 0; box-shadow: 0 10px 24px rgba(0,0,0,.25); }}
    .card h3 {{ margin: 0 0 8px 0; font-size: 14px; color: var(--muted); font-weight: 600; }}
    table {{ width: 100%; border-collapse: collapse; background: #151515; border: 1px solid var(--stroke); border-radius: 14px; overflow: hidden; }}
    th, td {{ border-bottom: 1px solid var(--stroke); padding: 12px; text-align: left; vertical-align: top; }}
    td:nth-child(1), td:nth-child(2) {{ white-space: nowrap; }}
    th {{ background: linear-gradient(180deg, #1e1e1e, #1a1a1a); font-weight: 700; color: var(--muted); position: static; top: auto; z-index: auto; }}
    tbody tr:nth-child(even) {{ background: #141414; }}
    tbody tr:hover {{ background: #1a1a1a; transition: background .15s ease; }}
    .score-badge {{ display: inline-block; padding: 2px 8px; border-radius: 999px; color: #111; font-size: 12px; font-weight: 700; background: var(--ti-orange-light); border: 1px solid var(--ti-orange-dark); }}
    .score-critical {{ background: #ff7043; border-color: #e64a19; }}
    .score-high {{ background: #ff8a65; border-color: #f4511e; }}
    .score-medium {{ background: #ffb74d; border-color: #fb8c00; }}
    .score-low {{ background: #ffd54f; border-color: #ffa000; }}
    .score-unknown {{ background: #9e9e9e; border-color: #757575; }}
    .cve-link {{ color: var(--ti-orange-light); text-decoration: none; font-weight: 600; }}
    .data-container {{ display: grid; gap: 6px; }}
    .data-preview {{ color: var(--text); line-height: 1.5; background: #121212; border: 1px solid var(--stroke); border-radius: 10px; padding: 10px; }}
    .expand-btn {{ border: 1px solid var(--stroke); background: linear-gradient(180deg, #1f1f1f, #181818); color: var(--text); padding: 8px 12px; border-radius: 10px; cursor: pointer; font-size: 12px; width: fit-content; transition: all .18s ease; }}
    .expand-btn:hover {{ border-color: var(--ti-orange); color: var(--ti-orange-light); transform: translateY(-1px); }}
    .json-data {{ white-space: pre-wrap; word-break: break-word; }}
    .footer {{ color: var(--muted); font-size: 12px; padding: 18px 0 6px; text-align: center; }}
    .legend-dot {{ display:inline-block; width:10px; height:10px; border-radius:50%; margin-right:6px; background: var(--ti-orange); }}
    /* Subpáginas de Detalhes */
    .detail-nav {{ display:flex; gap:10px; flex-wrap:wrap; margin: 6px 0 12px; }}
    .subtab-btn {{ border:1px solid var(--stroke); background: linear-gradient(180deg, #1f1f1f, #181818); color:var(--text); padding:8px 12px; border-radius:999px; cursor:pointer; font-size:12px; transition: all .18s ease; }}
    .subtab-btn:hover {{ border-color: var(--ti-orange); color: var(--ti-orange-light); transform: translateY(-1px); }}
    .subtab-btn.active {{ background: linear-gradient(180deg, #ffa726, #fb8c00); color:#111; border-color: var(--accent); font-weight:800; box-shadow: 0 6px 22px rgba(251,140,0,.25); }}
    .detail-subpage {{ display:none; }}
    .pager {{ display:flex; align-items:center; gap:8px; margin-top:8px; color: var(--muted); }}
    .pager button {{ border:1px solid var(--stroke); background:#181818; color:var(--text); padding:4px 8px; border-radius:6px; cursor:pointer; font-size:12px; }}
    .pager button:hover {{ border-color: var(--ti-orange); color: var(--ti-orange-light); }}
  </style>
  <script>
    function showTab(id) {{
      document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
      document.querySelectorAll('.tab-content').forEach(c => c.classList.remove('active'));
      document.getElementById('tab-btn-'+id).classList.add('active');
      document.getElementById('tab-'+id).classList.add('active');
    }}

    function toggleExpand(prefix, i) {{
      const prev = document.getElementById('preview-'+prefix+'-'+i);
      const full = document.getElementById('full-'+prefix+'-'+i);
      const btn = document.getElementById('btn-'+prefix+'-'+i);
      if (!prev || !full || !btn) return;
      if (full.style.display === 'none' || full.style.display === '') {{
        full.style.display = 'block';
        prev.style.display = 'none';
        btn.innerHTML = '<i class="expand-icon">▲</i> VER MENOS';
      }} else {{
        full.style.display = 'none';
        prev.style.display = 'block';
        btn.innerHTML = '<i class="expand-icon">▼</i> VER MAIS';
      }}
    }}

    function slugify(text) {{
      return (text || '').toString().toLowerCase().replace(/[^a-z0-9]+/gi,'-').replace(/^-+|-+$/g,'');
    }}

    function showDetail(slug) {{
      document.querySelectorAll('.subtab-btn').forEach(b => b.classList.remove('active'));
      document.querySelectorAll('.detail-subpage').forEach(s => s.style.display = 'none');
      const btn = document.getElementById('subtab-'+slug);
      const page = document.getElementById('detail-'+slug);
      if (btn) btn.classList.add('active');
      if (!page) return;
      page.style.display = 'block';
      initDetailPagination(slug);
      applyDetailPage(slug);
    }}

    function initDetailPagination(slug) {{
      const page = document.getElementById('detail-'+slug);
      if (!page) return;
      const sizeAttr = page.getAttribute('data-page-size');
      const pageSize = sizeAttr ? parseInt(sizeAttr, 10) : 50;
      page.setAttribute('data-page-size', pageSize);
      const rows = page.querySelectorAll('tbody tr');
      const total = rows.length;
      const pages = Math.max(1, Math.ceil(total / pageSize));
      page.setAttribute('data-pages', pages);
      if (!page.getAttribute('data-page')) page.setAttribute('data-page', '1');
      const curr = Math.min(pages, Math.max(1, parseInt(page.getAttribute('data-page')||'1',10)));
      page.setAttribute('data-page', String(curr));
      const pageEl = document.getElementById('detail-page-'+slug);
      const pagesEl = document.getElementById('detail-pages-'+slug);
      if (pageEl) pageEl.textContent = String(curr);
      if (pagesEl) pagesEl.textContent = String(pages);
    }}

    function applyDetailPage(slug) {{
      const page = document.getElementById('detail-'+slug);
      if (!page) return;
      const pageSize = parseInt(page.getAttribute('data-page-size')||'50',10);
      const curr = parseInt(page.getAttribute('data-page')||'1',10);
      const start = (curr-1)*pageSize;
      const end = start + pageSize;
      const rows = page.querySelectorAll('tbody tr');
      rows.forEach((r, idx) => {{ r.style.display = (idx>=start && idx<end) ? '' : 'none'; }});
      const pageEl = document.getElementById('detail-page-'+slug);
      if (pageEl) pageEl.textContent = String(curr);
    }}

    function changeDetailPage(slug, delta) {{
      const page = document.getElementById('detail-'+slug);
      if (!page) return;
      const pages = parseInt(page.getAttribute('data-pages')||'1',10);
      let curr = parseInt(page.getAttribute('data-page')||'1',10);
      curr = Math.min(pages, Math.max(1, curr + delta));
      page.setAttribute('data-page', String(curr));
      applyDetailPage(slug);
    }}

    // Desenho simples de gráfico de barras em Canvas (sem dependências)
    function drawBars(canvasId, labels, values) {{
      const canvas = document.getElementById(canvasId);
      if (!canvas) return;
      const ctx = canvas.getContext('2d');
      const dpr = window.devicePixelRatio || 1;
      const W = canvas.clientWidth * dpr;
      const H = canvas.clientHeight * dpr;
      canvas.width = W; canvas.height = H;
      ctx.scale(dpr, dpr);

      const pad = 28;
      const chartW = canvas.clientWidth - pad*2;
      const chartH = canvas.clientHeight - pad*2 - 20;
      const maxV = Math.max(1, ...values);
      const barGap = 10;
      const n = Math.max(1, values.length);
      const barW = Math.max(12, (chartW - barGap*(n-1)) / n);

      // Fundo
      ctx.fillStyle = '#141414';
      ctx.fillRect(0, 0, canvas.clientWidth, canvas.clientHeight);

      // Eixos
      ctx.strokeStyle = '#2a2a2a';
      ctx.beginPath();
      ctx.moveTo(pad, pad);
      ctx.lineTo(pad, pad + chartH);
      ctx.lineTo(pad + chartW, pad + chartH);
      ctx.stroke();

      // Barras
      const grad = ctx.createLinearGradient(0, pad, 0, pad + chartH);
      grad.addColorStop(0, '#ff9800');
      grad.addColorStop(1, '#e65100');

      for (let i=0;i<n;i++) {{
        const v = values[i] || 0;
        const h = (v / maxV) * (chartH - 2);
        const x = pad + i*(barW + barGap);
        const y = pad + chartH - h;
        ctx.fillStyle = grad;
        ctx.fillRect(x, y, barW, h);
      }}

      // Rótulos (limitados)
      ctx.fillStyle = '#bdbdbd';
      ctx.font = '12px Segoe UI, Arial';
      for (let i=0;i<n;i++) {{
        const label = (labels[i] || '').toString();
        const short = label.length > 18 ? label.slice(0,16) + '…' : label;
        const x = pad + i*(barW + barGap) + barW/2;
        const y = pad + chartH + 14;
        ctx.textAlign = 'center';
        ctx.fillText(short, x, y);
      }}

      // Título
      ctx.fillStyle = '#ff9800';
      ctx.font = 'bold 14px Segoe UI, Arial';
      ctx.textAlign = 'left';
      ctx.fillText('Top 10 tipos de eventos', pad, 20);
    }}

    window.addEventListener('DOMContentLoaded', function() {{
      try {{
        const labels = {labels_js};
        const values = {values_js};
        drawBars('eventsChart', labels, values);
      }} catch (e) {{}}
      try {{
        const detailSlugs = {detail_slugs_js};
        if (detailSlugs && detailSlugs.length) {{ showDetail(detailSlugs[0]); }}
      }} catch (e) {{}}
    }});

    function slugify(text) {{
      return (text || '').toString().toLowerCase().replace(/[^a-z0-9]+/gi,'-').replace(/^-+|-+$/g,'');
    }}

    function showDetail(slug) {{
      document.querySelectorAll('.subtab-btn').forEach(b => b.classList.remove('active'));
      document.querySelectorAll('.detail-subpage').forEach(s => s.style.display = 'none');
      const btn = document.getElementById('subtab-'+slug);
      const page = document.getElementById('detail-'+slug);
      if (btn) btn.classList.add('active');
      if (!page) return;
      page.style.display = 'block';
      initDetailPagination(slug);
      applyDetailPage(slug);
    }}

    function initDetailPagination(slug) {{
      const page = document.getElementById('detail-'+slug);
      if (!page) return;
      const sizeAttr = page.getAttribute('data-page-size');
      const pageSize = sizeAttr ? parseInt(sizeAttr, 10) : 50;
      page.setAttribute('data-page-size', pageSize);
      const rows = page.querySelectorAll('tbody tr');
      const total = rows.length;
      const pages = Math.max(1, Math.ceil(total / pageSize));
      page.setAttribute('data-pages', pages);
      if (!page.getAttribute('data-page')) page.setAttribute('data-page', '1');
      const curr = Math.min(pages, Math.max(1, parseInt(page.getAttribute('data-page')||'1',10)));
      page.setAttribute('data-page', String(curr));
      const pageEl = document.getElementById('detail-page-'+slug);
      const pagesEl = document.getElementById('detail-pages-'+slug);
      if (pageEl) pageEl.textContent = String(curr);
      if (pagesEl) pagesEl.textContent = String(pages);
    }}

    function applyDetailPage(slug) {{
      const page = document.getElementById('detail-'+slug);
      if (!page) return;
      const pageSize = parseInt(page.getAttribute('data-page-size')||'50',10);
      const curr = parseInt(page.getAttribute('data-page')||'1',10);
      const start = (curr-1)*pageSize;
      const end = start + pageSize;
      const rows = page.querySelectorAll('tbody tr');
      rows.forEach((r, idx) => {{ r.style.display = (idx>=start && idx<end) ? '' : 'none'; }});
      const pageEl = document.getElementById('detail-page-'+slug);
      if (pageEl) pageEl.textContent = String(curr);
    }}

    function changeDetailPage(slug, delta) {{
      const page = document.getElementById('detail-'+slug);
      if (!page) return;
      const pages = parseInt(page.getAttribute('data-pages')||'1',10);
      let curr = parseInt(page.getAttribute('data-page')||'1',10);
      curr = Math.min(pages, Math.max(1, curr + delta));
      page.setAttribute('data-page', String(curr));
      applyDetailPage(slug);
    }}
  </script>
  </head>
  <body>
    <header>
      <svg class="ti-lock" viewBox="0 0 24 24" aria-hidden="true">
        <path d="M12 1a5 5 0 00-5 5v3H6a2 2 0 00-2 2v8a2 2 0 002 2h12a2 2 0 002-2v-8a2 2 0 00-2-2h-1V6a5 5 0 00-5-5zm-3 8V6a3 3 0 016 0v3H9zm3 4a2 2 0 110 4 2 2 0 010-4z" fill="#ffffff"/>
      </svg>
      <h1>TI Safe Report — {html.escape(scan_name or 'Scan')} — Alvo: {html.escape(scan_target or '')}</h1>
    </header>
    <nav class="tabs">
      <button id="tab-btn-resumo" class="tab-btn active" onclick="showTab('resumo')">Resumo</button>
      <button id="tab-btn-vuln" class="tab-btn" onclick="showTab('vuln')">Vulnerabilidades</button>
      <button id="tab-btn-detalhes" class="tab-btn" onclick="showTab('detalhes')">Detalhes</button>
    </nav>
    <section id="tab-resumo" class="tab-content active">
      {resumo_html}
      <div class="card">
        <h3><span class="legend-dot"></span> Distribuição dos eventos (Top 10)</h3>
        <div style="width:100%; height:280px;">
          <canvas id="eventsChart" style="width:100%; height:260px;"></canvas>
        </div>
      </div>
    </section>
    <section id="tab-vuln" class="tab-content">{vulns_html}</section>
    <section id="tab-detalhes" class="tab-content">{detalhes_html}</section>
    <div class="footer">Relatório personalizado TI Safe — gerado por SpiderFoot</div>
  </body>
</html>
"""
        return page

    # ---------- helpers: HTML blocks ----------
    def _build_resumo_section(self, kpi: dict, total_events: int, scan_name: str, scan_target: str) -> str:
        return f"""
<div class="kpis">
  <div class="kpi"><div class="kpi-title">Eventos</div><div class="kpi-value">{total_events}</div></div>
  <div class="kpi"><div class="kpi-title">CVE Critical</div><div class="kpi-value">{kpi.get('CRITICAL',0)}</div></div>
  <div class="kpi"><div class="kpi-title">CVE High</div><div class="kpi-value">{kpi.get('HIGH',0)}</div></div>
  <div class="kpi"><div class="kpi-title">CVE Medium</div><div class="kpi-value">{kpi.get('MEDIUM',0)}</div></div>
  <div class="kpi"><div class="kpi-title">CVE Low</div><div class="kpi-value">{kpi.get('LOW',0)}</div></div>
</div>
"""

    def _build_cve_section(self, events: list) -> str:
        # Filtrar somente eventos de vulnerabilidade
        cve_events = [e for e in events if e.get("event_type") in (
            "VULNERABILITY_CVE_CRITICAL", "VULNERABILITY_CVE_HIGH", "VULNERABILITY_CVE_MEDIUM", "VULNERABILITY_CVE_LOW", "VULNERABILITY_GENERAL"
        )]

        # Agrupar por CVE ID (eventos do Shodan costumam repetir o mesmo CVE)
        grouped = {}
        for e in cve_events:
            cve_id, score, description = self.parse_cve_data(e.get("data", ""))
            key = cve_id if cve_id and cve_id != 'N/A' else f"GENERAL:{e.get('event_type','')}"
            grouped.setdefault(key, {"events": [], "best_score": score, "description": description})
            grouped[key]["events"].append(e)
            # Escolher o maior score numérico disponível como representativo
            try:
                curr = float(str(grouped[key]["best_score"])) if grouped[key]["best_score"] not in (None, "Unknown") else -1.0
            except Exception:
                curr = -1.0
            try:
                cand = float(str(score)) if score not in (None, "Unknown") else -1.0
            except Exception:
                cand = -1.0
            if cand > curr:
                grouped[key]["best_score"] = score
            # Preferir primeira descrição não vazia
            if (not grouped[key]["description"] or grouped[key]["description"] == "Unknown") and description and isinstance(description, str):
                grouped[key]["description"] = description

        rows = []
        idx = 0
        for key, payload in grouped.items():
            # Derivar colunas agregadas
            cve_id = key if not key.startswith("GENERAL:") else "VULNERABILITY_GENERAL"
            mitre_link = f"https://cve.mitre.org/cgi-bin/cvename.cgi?name={html.escape(cve_id)}" if cve_id.startswith("CVE-") else None
            cve_cell = f"<a class=\"cve-link\" target=\"_blank\" rel=\"noopener\" href=\"{mitre_link}\">{html.escape(cve_id)}</a>" if mitre_link else html.escape(cve_id)

            score = payload.get("best_score", "Unknown")
            score_class = self.get_score_class(score)
            score_badge = f"<span class=\"score-badge {score_class}\">{html.escape(str(score))}</span>"

            description = payload.get("description", "Unknown") or "Unknown"
            desc_preview = (description[:160] + "...") if isinstance(description, str) and len(description) > 160 else description

            # Data/status agregados (mais recente e se há algum FP)
            last_dates = [e.get("last_seen", "") for e in payload["events"]]
            date_str = html.escape(max(last_dates) if last_dates else "")
            any_fp = any(str(e.get("false_positive", "0")) == "1" for e in payload["events"])
            status_str = "FP" if any_fp else "OK"

            # Tabela de ocorrências para expandir
            occurrences = []
            for j, ev in enumerate(payload["events"]):
                occurrences.append(
                    f"<tr><td>{html.escape(ev.get('module',''))}</td><td>{html.escape(ev.get('last_seen',''))}</td><td>{'FP' if str(ev.get('false_positive','0'))=='1' else 'OK'}</td></tr>"
                )
            occ_table = (
                "<div class=\"data-content\">"
                "<table style=\"width:100%; border-collapse:collapse;\"><thead>"
                "<tr><th>Módulo</th><th>Data</th><th>Status</th></tr></thead><tbody>"
                + "".join(occurrences) + "</tbody></table></div>"
            )

            row = f"""
<tr>
  <td>{cve_cell} <span style=\"margin-left:6px; color:#bdbdbd; font-size:12px;\">({len(payload['events'])})</span></td>
  <td>{score_badge}</td>
  <td>
    <div class=\"data-container\">
      <div class=\"data-preview\" id=\"preview-CVE-{idx}\"><div class=\"description-content\">{html.escape(str(desc_preview or ''))}</div></div>
      <button class=\"expand-btn\" onclick=\"toggleExpand('CVE', {idx})\" id=\"btn-CVE-{idx}\"><i class=\"expand-icon\">▼</i> VER MAIS</button>
      <div class=\"data-full\" id=\"full-CVE-{idx}\" style=\"display:none;\"><div class=\"data-content\"><div class=\"json-data\">{html.escape(str(description or ''))}</div></div>{occ_table}</div>
    </div>
  </td>
  <td>{date_str}</td>
  <td>{status_str}</td>
</tr>
"""
            rows.append(row)
            idx += 1

        table = f"""
<h3>Vulnerabilidades (CVEs)</h3>
<table>
  <thead>
    <tr>
      <th>CVE</th>
      <th>Score</th>
      <th>Descrição</th>
      <th>Data</th>
      <th>Status</th>
    </tr>
  </thead>
  <tbody>
    {''.join(rows) if rows else '<tr><td colspan="5">Nenhuma CVE encontrada.</td></tr>'}
  </tbody>
 </table>
"""
        return table

    def _build_details_section(self, events: list):
        # Agrupar por tipo (tópico)
        by_type = {}
        for e in events:
            et = e.get("event_type", "UNKNOWN")
            by_type.setdefault(et, []).append(e)

        if not by_type:
            return "<div>Nenhum dado.</div>", []

        nav_parts = ["<div class=\"detail-nav\">"]
        subpages_parts = []
        slugs = []

        def mk_slug(name: str) -> str:
            import re as _re
            s = _re.sub(r"[^a-zA-Z0-9]+", "-", name or "UNKNOWN").strip("-")
            return s.lower() or "unknown"

        # Navbar de subtópicos
        for etype in sorted(by_type.keys()):
            slug = mk_slug(etype)
            slugs.append(slug)
            nav_parts.append(
                f"<button id=\"subtab-{slug}\" class=\"subtab-btn\" onclick=\"showDetail('{slug}')\">{html.escape(etype)}</button>"
            )
        nav_parts.append("</div>")

        # Subpáginas
        for etype in sorted(by_type.keys()):
            slug = mk_slug(etype)
            rows_html = []
            for i, e in enumerate(by_type[etype]):
                data = e.get("data", "")
                prev = (data[:160] + "...") if isinstance(data, str) and len(data) > 160 else data
                status_str = "FP" if str(e.get("false_positive", "0")) == "1" else "OK"
                rows_html.append(
                    f"""
<tr>
  <td>{html.escape(etype)}</td>
  <td>{html.escape(e.get('module',''))}</td>
  <td>
    <div class=\"data-container\">\n      <div class=\"data-preview\" id=\"preview-DET-{etype}-{i}\">{html.escape(str(prev or ''))}</div>
      <button class=\"expand-btn\" onclick=\"toggleExpand('DET-{etype}', {i})\" id=\"btn-DET-{etype}-{i}\"><i class=\"expand-icon\">▼</i> VER MAIS</button>
      <div class=\"data-full\" id=\"full-DET-{etype}-{i}\" style=\"display:none;\"><div class=\"json-data\">{html.escape(str(data or ''))}</div></div>
    </div>
  </td>
  <td>{html.escape(e.get('last_seen',''))}</td>
  <td>{status_str}</td>
</tr>
"""
                )

            subpages_parts.append(
                f"""
<div class=\"detail-subpage\" id=\"detail-{slug}\" data-page-size=\"50\" data-page=\"1\">
  <table>
    <thead><tr><th>Tipo</th><th>Módulo</th><th>Conteúdo</th><th>Última vez</th><th>Status</th></tr></thead>
    <tbody>
      {''.join(rows_html)}
    </tbody>
  </table>
  <div class=\"pager\">
    <button onclick=\"changeDetailPage('{slug}', -1)\">Anterior</button>
    <span>Página <span id=\"detail-page-{slug}\">1</span> de <span id=\"detail-pages-{slug}\">1</span></span>
    <button onclick=\"changeDetailPage('{slug}', 1)\">Próxima</button>
  </div>
</div>
"""
            )

        html_block = "".join(nav_parts + subpages_parts)
        return html_block, slugs

    # ---------- helpers: CVE parsing ----------
    def parse_cve_data(self, data: str):
        """Extrai CVE ID, score e descrição.

        Suporta:
        - Texto simples (CVE-YYYY-NNNN, "Score:", "Description:")
        - JSON (campos: cve/cve_id/id, cvss_v3/cvss_v2/cvss/score, summary/description)
        """
        import re, json
        if not data:
            return 'N/A', 'Unknown', 'Unknown'

        cve_id = 'N/A'
        score = 'Unknown'
        description = 'Unknown'

        # 1) Tentar interpretar como JSON primeiro
        try:
            s = data if isinstance(data, str) else str(data)
            st = s.strip()
            obj = None
            if st.startswith('{') or st.startswith('['):
                obj = json.loads(st)
                if isinstance(obj, list) and obj:
                    obj = obj[0]
                if isinstance(obj, dict):
                    low = {str(k).lower(): v for k, v in obj.items()}
                    # ID
                    for k in ('cve', 'cve_id', 'id', 'cveid'):
                        if k in low and isinstance(low[k], (str, int)):
                            cve_id = str(low[k]).strip()
                            break
                    # SCORE
                    for k in ('cvss_v3', 'cvss3', 'cvss_v2', 'cvss2', 'cvss', 'score'):
                        if k in low and low[k] is not None:
                            try:
                                score = str(float(low[k]))
                            except Exception:
                                score = str(low[k])
                            break
                    # DESCRIPTION
                    for k in ('summary', 'description', 'desc'):
                        if k in low and isinstance(low[k], str) and low[k].strip():
                            description = low[k].strip()
                            break
        except Exception:
            pass

        # 2) Se ainda faltarem campos, tentar parsing de texto
        if cve_id == 'N/A' or score == 'Unknown' or description == 'Unknown':
            lines = str(data).split("\n")
            if cve_id == 'N/A':
                for line in lines:
                    line = line.strip()
                    if line.startswith('CVE-'):
                        cve_id = line
                        break
                if cve_id == 'N/A':
                    for line in lines:
                        line = line.strip()
                        m = re.search(r'CVE-\d{4}-\d+', line)
                        if m:
                            cve_id = m.group(0)
                            break
            if score == 'Unknown':
                for line in lines:
                    line = line.strip()
                    if line.lower().startswith('score:'):
                        score_part = line.split(':', 1)[1].strip()
                        m = re.search(r'(\d+\.?\d*)', score_part)
                        if m:
                            score = m.group(1)
                        break
            if description == 'Unknown':
                for line in lines:
                    line = line.strip()
                    if line.lower().startswith('description:'):
                        description = line.split(':', 1)[1].strip()
                        break

        if description == 'Unknown' and cve_id != 'N/A':
            description = f"Vulnerabilidade de segurança identificada: {cve_id}"

        return cve_id, score, description

    def get_score_class(self, score: str) -> str:
        if score == 'Unknown':
            return 'score-unknown'
        try:
            sn = float(score)
            if sn >= 9.0:
                return 'score-critical'
            if sn >= 7.0:
                return 'score-high'
            if sn >= 4.0:
                return 'score-medium'
            return 'score-low'
        except Exception:
            return 'score-unknown'
